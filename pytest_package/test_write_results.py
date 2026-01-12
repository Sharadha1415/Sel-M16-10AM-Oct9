## Stored the login credentials in an excel file(ddt\login_credentials.xlsx)
## reading the login credentials(ddt\read_login_data.py)

import os.path
import time
from openpyxl import Workbook
import pytest
from openpyxl.reader.excel import load_workbook

from selenium import webdriver
from ddt.read_login_data import read_login_credentials

opts = webdriver.ChromeOptions()
opts.add_experimental_option("detach", True)

login_data = read_login_credentials()

def write_excel(test_name, status):
    if not os.path.exists("results.xlsx"):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "test_results"
        sheet.append(["Test Name", "Status"])
        wb.save("results.xlsx")

    wb = load_workbook("results.xlsx")
    sheet = wb["test_results"]

    sheet.append([test_name, status])
    wb.save("results.xlsx")


@pytest.fixture()
def setup():
    driver = webdriver.Chrome(opts)
    driver.get('https://www.saucedemo.com/')
    time.sleep(2)
    yield driver
    driver.close()

@pytest.mark.parametrize("username, pwd", login_data)
def test_login(username, pwd, setup):
    setup.find_element("id", "user-name").send_keys(username)
    time.sleep(1)
    setup.find_element('id', 'password').send_keys(pwd)
    time.sleep(1)
    setup.find_element('id', 'login-button').click()
    time.sleep(3)

    try:
        assert 'inventory' in setup.current_url
        print("successfull login")
        write_excel(f'{username}{pwd}', "passed")
    except:
        print("unsuccessfull login")
        write_excel(f'{username}{pwd}', "failed")

